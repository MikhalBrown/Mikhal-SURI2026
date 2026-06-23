"""
SensorTOOL Data to Excel Exporter
Converts sensorTOOL 2.3.1.4177 data exports to organized Excel files
with labeled columns for thermal and spectral analysis.

Author: Data Processing Pipeline
Date: 2026
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import sys
from pathlib import Path
from typing import Dict, List, Optional


class SensorToolExporter:
    """
    Handles conversion of sensorTOOL exported data to organized Excel files.
    
    Target columns:
    - Peak temperature
    - Mean temperature
    - Cooling rate
    - Temperature variance
    - Thermal gradient
    - Peak frequency
    - Spectral energy
    """
    
    def __init__(self, input_file: str, output_file: Optional[str] = None):
        """
        Initialize the exporter.
        
        Args:
            input_file: Path to sensorTOOL exported file (CSV, JSON, or Excel)
            output_file: Path for output Excel file (default: input_file_formatted.xlsx)
        """
        self.input_file = input_file
        self.output_file = output_file or self._generate_output_filename(input_file)
        
        # Define target columns
        self.target_columns = [
            'Peak Temperature (°C)',
            'Mean Temperature (°C)',
            'Cooling Rate (°C/s)',
            'Temperature Variance (°C²)',
            'Thermal Gradient (°C/mm)',
            'Peak Frequency (Hz)',
            'Spectral Energy (dB)'
        ]
        
        # Common column name mappings from sensorTOOL
        self.column_mappings = {
            'Peak Temperature': ['peak temp', 'peak temperature', 'peak_temperature', 'max temp', 'maximum temperature', 'T_peak', 'temp_peak'],
            'Mean Temperature': ['mean temp', 'mean temperature', 'mean_temperature', 'avg temp', 'average temperature', 'T_mean', 'temp_mean'],
            'Cooling Rate': ['cooling rate', 'cool rate', 'cooling_rate', 'cool_rate', 'rate of cooling'],
            'Temperature Variance': ['temp variance', 'temperature variance', 'temperature_variance', 'temp_var', 'variance', 'T_variance'],
            'Thermal Gradient': ['thermal gradient', 'temp gradient', 'gradient', 'thermal_gradient', 'T_gradient'],
            'Peak Frequency': ['peak freq', 'peak frequency', 'peak_frequency', 'freq', 'frequency', 'peak_freq', 'f_peak'],
            'Spectral Energy': ['spectral energy', 'energy', 'spectral_energy', 'energy_spectral', 'S_energy']
        }
    
    @staticmethod
    def _generate_output_filename(input_file: str) -> str:
        """Generate output filename based on input file."""
        base = Path(input_file).stem
        return f"{base}_formatted.xlsx"
    
    def _read_data(self) -> pd.DataFrame:
        """
        Read data from input file.
        
        Returns:
            DataFrame with raw data from sensorTOOL export
        """
        file_ext = Path(self.input_file).suffix.lower()
        
        if file_ext == '.csv':
            # Try different encodings and delimiters for CSV
            try:
                df = pd.read_csv(self.input_file, encoding='utf-8')
            except UnicodeDecodeError:
                df = pd.read_csv(self.input_file, encoding='latin-1')
        
        elif file_ext in ['.xlsx', '.xls']:
            df = pd.read_excel(self.input_file)
        
        elif file_ext == '.json':
            df = pd.read_json(self.input_file)
        
        else:
            raise ValueError(f"Unsupported file format: {file_ext}. Use CSV, JSON, or Excel.")
        
        return df
    
    def _find_column_mapping(self, df_columns: List[str]) -> Dict[str, str]:
        """
        Find mapping between input columns and target columns.
        
        Args:
            df_columns: List of column names from input data
            
        Returns:
            Dictionary mapping target columns to input columns
        """
        mapping = {}
        df_cols_lower = [col.lower().strip() for col in df_columns]
        
        for target, aliases in self.column_mappings.items():
            for alias in aliases:
                for df_col_lower, df_col_original in zip(df_cols_lower, df_columns):
                    if alias.lower() in df_col_lower or df_col_lower in alias.lower():
                        mapping[target] = df_col_original
                        break
                if target in mapping:
                    break
        
        return mapping
    
    def _process_data(self, df: pd.DataFrame, column_mapping: Dict[str, str]) -> pd.DataFrame:
        """
        Process and organize the data.
        
        Args:
            df: Input DataFrame
            column_mapping: Mapping of target columns to source columns
            
        Returns:
            Processed DataFrame with target columns
        """
        # Extract only the mapped columns
        extracted_cols = {}
        missing_columns = []
        
        for target, source in column_mapping.items():
            if source in df.columns:
                extracted_cols[target] = df[source]
            else:
                missing_columns.append(target)
        
        # Log missing columns
        if missing_columns:
            print(f"Warning: The following columns were not found in input data:")
            for col in missing_columns:
                print(f"  - {col}")
        
        # Create new DataFrame with organized columns
        result_df = pd.DataFrame(extracted_cols)
        
        # Convert columns to numeric, handling errors
        for col in result_df.columns:
            result_df[col] = pd.to_numeric(result_df[col], errors='coerce')
        
        # Add units to column names
        result_df = result_df.rename(columns={
            'Peak Temperature': 'Peak Temperature (°C)',
            'Mean Temperature': 'Mean Temperature (°C)',
            'Cooling Rate': 'Cooling Rate (°C/s)',
            'Temperature Variance': 'Temperature Variance (°C²)',
            'Thermal Gradient': 'Thermal Gradient (°C/mm)',
            'Peak Frequency': 'Peak Frequency (Hz)',
            'Spectral Energy': 'Spectral Energy (dB)'
        })
        
        return result_df
    
    def _format_excel(self, output_path: str):
        """
        Apply formatting to the Excel file.
        
        Args:
            output_path: Path to the Excel file to format
        """
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Border style
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting
        for cell in ws[1]:
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
        
        # Apply data formatting
        center_alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = center_alignment
                cell.border = border
                # Format numbers to 4 decimal places
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0000'
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
    
    def convert(self, auto_detect: bool = True, column_mapping: Optional[Dict[str, str]] = None):
        """
        Convert sensorTOOL data to formatted Excel file.
        
        Args:
            auto_detect: Automatically detect column mappings (default: True)
            column_mapping: Custom column mapping (overrides auto-detection)
        """
        try:
            print(f"Reading input file: {self.input_file}")
            df = self._read_data()
            print(f"Input file contains {len(df)} rows and {len(df.columns)} columns")
            print(f"Columns found: {list(df.columns)}")
            
            # Determine column mapping
            if column_mapping is None:
                if auto_detect:
                    print("\nAttempting automatic column detection...")
                    column_mapping = self._find_column_mapping(df.columns)
                    print(f"Detected {len(column_mapping)} columns")
                else:
                    raise ValueError("No column mapping provided and auto_detect is False")
            
            print(f"\nColumn mapping:")
            for target, source in column_mapping.items():
                print(f"  {target} <- {source}")
            
            # Process data
            print("\nProcessing data...")
            processed_df = self._process_data(df, column_mapping)
            
            # Export to Excel
            print(f"\nExporting to Excel: {self.output_file}")
            processed_df.to_excel(self.output_file, sheet_name='Thermal Data', index=False)
            
            # Format Excel file
            print("Applying formatting...")
            self._format_excel(self.output_file)
            
            print(f"\n✓ Successfully exported to: {self.output_file}")
            print(f"Output contains {len(processed_df)} rows and {len(processed_df.columns)} columns")
            
            return processed_df
        
        except FileNotFoundError:
            print(f"Error: Input file not found: {self.input_file}")
            sys.exit(1)
        except Exception as e:
            print(f"Error during conversion: {str(e)}")
            sys.exit(1)


def main():
    """Main entry point for the script."""
    
    # Example usage
    if len(sys.argv) < 2:
        print("Usage: python sensortool_to_excel.py <input_file> [output_file]")
        print("\nExample:")
        print("  python sensortool_to_excel.py data.csv")
        print("  python sensortool_to_excel.py thermal_data.xlsx output_formatted.xlsx")
        sys.exit(0)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    # Create exporter and convert
    exporter = SensorToolExporter(input_file, output_file)
    exporter.convert()


if __name__ == "__main__":
    main()
